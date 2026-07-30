from __future__ import annotations

import csv
import hashlib
from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Any, Iterable

from sqlalchemy import select
from sqlalchemy.orm import Session

from leopard_project.config import load_seed_bundle, normalize_alias
from leopard_project.market_paths import load_market_path_registry, market_path_mapping
from leopard_project.models import DailyBar, DataStatus, LiquidityStatus, Market
from leopard_project.providers import ProviderError, ThsPublicValidationProvider
from leopard_project.providers.capabilities import load_provider_capabilities

from .enhanced import calculate_market_metrics
from .models import MarketRefreshItem, MarketRefreshRun, SectorDailyBar, SectorIndicatorSnapshot
from .services import WebDomainError
from .write_coordination import BACKGROUND_WRITE_LOCK


PROVIDER_KEY = "ths_public_validation"
PROVIDER_ROLE = "diagnostic_provider"
BAR_COMPARE_QUANTUM = Decimal("0.000001")


def _bar_signature(close, pre_close, pct_change) -> tuple[Decimal, Decimal, Decimal]:
    return tuple(Decimal(str(value)).quantize(BAR_COMPARE_QUANTUM) for value in (close, pre_close, pct_change))  # type: ignore[return-value]


def _liquidity(bar: DailyBar) -> str:
    return bar.liquidity_status.value if hasattr(bar.liquidity_status, "value") else str(bar.liquidity_status)


def _persist_bar(session: Session, sector_key: str, bar: DailyBar, source: str, role: str) -> str:
    existing = session.scalar(select(SectorDailyBar).where(
        SectorDailyBar.sector_key == sector_key,
        SectorDailyBar.trade_date == bar.trade_date,
    ))
    values = {
        "open": bar.open, "high": bar.high, "low": bar.low, "close": bar.close,
        "pre_close": bar.pre_close, "daily_pct_change": bar.pct_change,
        "volume": bar.volume, "amount": bar.amount, "turnover_rate": bar.turnover_rate,
        "liquidity_status": _liquidity(bar), "eod_status": "complete_eod",
        "data_source": source, "provider_role": role, "fetched_at": bar.fetched_at,
        "source_response_hash": bar.source_payload_hash,
    }
    if existing:
        comparable = _bar_signature(existing.close, existing.pre_close, existing.daily_pct_change)
        incoming = _bar_signature(bar.close, bar.pre_close, bar.pct_change)
        if comparable != incoming:
            return "conflict"
        return "duplicate"
    session.add(SectorDailyBar(sector_key=sector_key, trade_date=bar.trade_date, **values))
    return "inserted"


def _recalculate(session: Session, sector_key: str) -> None:
    session.flush()
    bars = list(session.scalars(select(SectorDailyBar).where(
        SectorDailyBar.sector_key == sector_key,
        SectorDailyBar.eod_status == "complete_eod",
    ).order_by(SectorDailyBar.trade_date)))
    for index, bar in enumerate(bars, start=1):
        metrics = calculate_market_metrics(bars[:index])
        indicator = session.scalar(select(SectorIndicatorSnapshot).where(SectorIndicatorSnapshot.daily_bar_id == bar.id))
        if indicator is None:
            session.add(SectorIndicatorSnapshot(daily_bar_id=bar.id, **metrics))


def _composite_bars(symbol: str, components: list[list[DailyBar]]) -> list[DailyBar]:
    by_component = [{bar.trade_date: bar for bar in rows} for rows in components]
    common_dates = sorted(set.intersection(*(set(rows) for rows in by_component)))
    result: list[DailyBar] = []
    close = Decimal("1000")
    for day in common_dates:
        rows = [values[day] for values in by_component]
        returns = [(bar.close / bar.pre_close - 1) if bar.pre_close else Decimal("0") for bar in rows]
        daily_return = sum(returns, Decimal("0")) / Decimal(len(returns))
        pre_close = close
        close = pre_close * (Decimal("1") + daily_return)
        open_ratio = sum((bar.open / bar.pre_close for bar in rows), Decimal("0")) / Decimal(len(rows))
        high_ratio = sum((bar.high / bar.pre_close for bar in rows), Decimal("0")) / Decimal(len(rows))
        low_ratio = sum((bar.low / bar.pre_close for bar in rows), Decimal("0")) / Decimal(len(rows))
        volume = sum((bar.volume or Decimal("0") for bar in rows), Decimal("0")) or None
        amount = sum((bar.amount or Decimal("0") for bar in rows), Decimal("0")) or None
        digest = hashlib.sha256(":".join(bar.source_payload_hash for bar in rows).encode()).hexdigest()
        result.append(DailyBar(
            symbol=symbol, symbol_name=symbol, market=Market.CN_A, trade_date=day,
            open=pre_close * open_ratio, high=pre_close * high_ratio, low=pre_close * low_ratio,
            close=close, pre_close=pre_close, change=close - pre_close,
            pct_change=daily_return * Decimal("100"), volume=volume, amount=amount,
            turnover_rate=None,
            liquidity_status=LiquidityStatus.COMPLETE if volume is not None and amount is not None else LiquidityStatus.PARTIAL,
            provider=PROVIDER_KEY, fetched_at=max(bar.fetched_at for bar in rows),
            source_payload_hash=digest, data_status=DataStatus.NORMAL,
        ))
    return result


def refresh_real_market(
    session: Session,
    actor: str,
    *,
    sector_keys: list[str] | None,
    as_of: date,
    provider: ThsPublicValidationProvider | None = None,
    mode: str = "manual_real_refresh",
    allowed_trade_dates: set[date] | None = None,
    allowed_trade_dates_by_sector: dict[str, set[date]] | None = None,
    attempt_number: int = 1,
    next_retry_at: datetime | None = None,
) -> MarketRefreshRun:
    """Fetch outside SQLite transactions, then persist each normalized sector briefly."""
    bundle = load_seed_bundle()
    registry = load_market_path_registry(bundle)
    capabilities = load_provider_capabilities()
    sectors = list(registry.supported_market_paths)
    mappings = {item.market_path_key: market_path_mapping(item, bundle) for item in sectors}
    if sector_keys:
        requested = set(sector_keys)
        sectors = [item for item in sectors if item.market_path_key in requested]
    client = provider or ThsPublicValidationProvider(minimum_interval=0.45)
    # The public endpoint is year-partitioned and some newly introduced boards
    # legitimately have no prior-year resource. The current year already
    # covers the Phase 2A-0 120-session contract at the accepted July cutoff.
    start = date(as_of.year, 1, 1)
    fetched: list[tuple[object, list[DailyBar] | None, str | None, str | None, str | None, str | None]] = []
    for sector in sectors:
        sector_key = sector.market_path_key
        mapping = mappings[sector_key]
        capability = capabilities[sector_key]
        selected = capability.selectable_candidates[0] if capability.selectable_candidates else None
        sector_allowed_dates = (
            allowed_trade_dates_by_sector.get(sector_key, set())
            if allowed_trade_dates_by_sector is not None
            else allowed_trade_dates
        )
        try:
            if selected is None:
                raise WebDomainError("provider_unverified", f"{sector.display_name}没有已验证运行链", 422)
            proxy = sector.mapping_type == "proxy"
            symbols = [str(item["symbol"]) for item in selected.components] if selected.components else [selected.symbol]
            component_rows = [list(client.historical_daily_bars(symbol, start, as_of, Market.CN_A)) for symbol in symbols]
            if not component_rows or any(not rows for rows in component_rows):
                raise WebDomainError("market_no_data", f"{sector.display_name}没有可用行情", 422)
            bars = _composite_bars(selected.symbol, component_rows) if len(component_rows) > 1 else component_rows[0]
            source = f"{PROVIDER_KEY}:proxy:{selected.symbol}" if proxy else f"{PROVIDER_KEY}:custom_composite" if len(component_rows) > 1 else PROVIDER_KEY
            eligible = [
                bar for bar in bars
                if bar.trade_date <= as_of
                and bar.data_status == DataStatus.NORMAL
                and (sector_allowed_dates is None or bar.trade_date in sector_allowed_dates)
            ]
            if sector_allowed_dates is not None and not eligible:
                raise WebDomainError("market_missing_dates_not_returned", "Provider未返回请求的缺失交易日", 422)
            fetched.append((sector, eligible, source, "short_history" if len(bars) < 21 else "complete_eod", None, None))
        except (ProviderError, WebDomainError, ValueError) as exc:
            code = exc.code if isinstance(exc, WebDomainError) else exc.category.value if isinstance(exc, ProviderError) else "validation_error"
            pending = sector_allowed_dates is not None and code in {"market_no_data", "market_missing_dates_not_returned", "no_data"}
            fetched.append((sector, None, None, "pending_publication" if pending else "provider_failed", code, "Provider daily data is not published yet" if pending else "Provider request failed"))

    attempted_at = datetime.now(timezone.utc)
    run = MarketRefreshRun(
        mode=mode,
        provider_role=PROVIDER_ROLE,
        requested_count=len(sectors),
        requested_by=actor,
        status="running",
    )
    with BACKGROUND_WRITE_LOCK:
        session.add(run)
        session.commit()
    for sector, eligible, source, status, error_code, error_message in fetched:
        try:
            with BACKGROUND_WRITE_LOCK:
                if eligible is None:
                    run.failure_count += 1
                    session.add(MarketRefreshItem(
                        run_id=run.id,
                        sector_key=sector.market_path_key,
                        status=status,
                        expected_trade_date=as_of,
                        attempt_number=attempt_number,
                        attempted_at=attempted_at,
                        next_retry_at=next_retry_at if status == "pending_publication" else None,
                        error_code=error_code,
                        error_message=error_message,
                        detail=status,
                    ))
                else:
                    conflicts = 0
                    for bar in eligible:
                        existing = session.scalar(select(SectorDailyBar).where(
                            SectorDailyBar.sector_key == sector.market_path_key,
                            SectorDailyBar.trade_date == bar.trade_date,
                        ))
                        if existing is not None:
                            conflicts += _bar_signature(existing.close, existing.pre_close, existing.daily_pct_change) != _bar_signature(bar.close, bar.pre_close, bar.pct_change)
                    if conflicts:
                        raise WebDomainError("market_data_conflict", f"{conflicts} conflicting rows", 409)
                    for bar in eligible:
                        _persist_bar(session, sector.market_path_key, bar, source or PROVIDER_KEY, PROVIDER_ROLE)
                    _recalculate(session, sector.market_path_key)
                    latest = max(bar.trade_date for bar in eligible)
                    run.short_history_count += status == "short_history"
                    run.success_count += 1
                    capability = capabilities[sector.market_path_key]
                    selected = capability.selectable_candidates[0]
                    proxy = sector.mapping_type == "proxy"
                    provider_symbol = selected.symbol
                    lineage = (
                        f"canonical_market_path={sector.market_path_key};parent_report_topic={sector.parent_report_topic};"
                        f"mapping_type=proxy;proxy_symbol={provider_symbol};provider={PROVIDER_KEY};"
                        f"provider_symbol={provider_symbol};provider_name={selected.provider_name};"
                        f"rationale=explicit tourism-and-hotel proxy;as_of={as_of.isoformat()};source_status=available"
                        if proxy else
                        f"provider={PROVIDER_KEY};provider_symbol={provider_symbol};"
                        f"as_of={as_of.isoformat()};source_status=available"
                    )
                    session.add(MarketRefreshItem(
                        run_id=run.id,
                        sector_key=sector.market_path_key,
                        status=status,
                        trade_date=latest,
                        expected_trade_date=as_of,
                        attempt_number=attempt_number,
                        attempted_at=attempted_at,
                        completed_at=datetime.now(timezone.utc),
                        provider=PROVIDER_KEY,
                        provider_symbol=provider_symbol,
                        lineage=lineage,
                        detail=f"{source}; {len(eligible)} eligible rows",
                    ))
                session.commit()
        except Exception:
            session.rollback()
            raise
    pending_count = sum(1 for item in fetched if item[3] == "pending_publication")
    with BACKGROUND_WRITE_LOCK:
        run.status = "pending_retry" if pending_count else "partial_failure" if run.failure_count else "complete"
        run.finished_at = datetime.now(timezone.utc)
        session.commit()
    return run


def _rows_from_upload(filename: str, payload: bytes) -> list[dict[str, Any]]:
    if filename.lower().endswith(".csv"):
        return list(csv.DictReader(StringIO(payload.decode("utf-8-sig"))))
    if filename.lower().endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise WebDomainError("xlsx_dependency_unavailable", "当前环境无法读取Excel，请改用CSV", 422) from exc
        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        return [dict(zip(headers, values)) for values in rows[1:]]
    raise WebDomainError("market_import_type", "只支持CSV或XLSX", 422)


def import_real_market(session: Session, actor: str, filename: str, payload: bytes, *, confirmed: bool) -> dict[str, Any]:
    rows = _rows_from_upload(filename, payload)
    required = {"trade_date", "open", "high", "low", "close", "pre_close", "volume"}
    preview: list[dict[str, Any]] = []
    normalized: list[tuple[str, DailyBar, str]] = []
    bundle = load_seed_bundle()
    valid_keys = {item.sector_key for item in bundle.sectors}
    file_hash = hashlib.sha256(payload).hexdigest()
    for number, row in enumerate(rows, start=2):
        missing = required - set(row)
        if missing:
            preview.append({"row": number, "status": "invalid", "detail": f"缺少字段：{','.join(sorted(missing))}"}); continue
        sector_key = str(row.get("sector_key") or "").strip() or normalize_alias(str(row.get("sector_name") or "").strip(), bundle)
        if sector_key not in valid_keys or sector_key == "hang_seng_tech":
            preview.append({"row": number, "status": "needs_confirmation", "detail": "板块无法映射"}); continue
        try:
            close, pre_close = Decimal(str(row["close"])), Decimal(str(row["pre_close"]))
            volume = Decimal(str(row["volume"])) if row["volume"] not in (None, "") else None
            amount = Decimal(str(row.get("amount"))) if row.get("amount") not in (None, "") else None
            turnover = Decimal(str(row.get("turnover_rate"))) if row.get("turnover_rate") not in (None, "") else None
            liquidity = LiquidityStatus.COMPLETE if volume is not None and amount is not None else LiquidityStatus.PARTIAL if any(v is not None for v in (volume, amount, turnover)) else LiquidityStatus.UNAVAILABLE
            source = str(row.get("source_name") or "manual_file_import").strip()
            bar = DailyBar(symbol=sector_key, symbol_name=sector_key, market=Market.CN_A,
                trade_date=date.fromisoformat(str(row["trade_date"])[:10]), open=Decimal(str(row["open"])),
                high=Decimal(str(row["high"])), low=Decimal(str(row["low"])), close=close, pre_close=pre_close,
                change=close-pre_close, pct_change=(close/pre_close-1)*100 if pre_close else Decimal("0"),
                volume=volume, amount=amount, turnover_rate=turnover, liquidity_status=liquidity,
                provider=source, fetched_at=datetime.now(timezone.utc),
                source_payload_hash=hashlib.sha256(f"{file_hash}:{number}".encode()).hexdigest(), data_status=DataStatus.NORMAL)
            normalized.append((sector_key, bar, source)); preview.append({"row": number, "sector_key": sector_key, "trade_date": bar.trade_date.isoformat(), "status": "ready"})
        except (ValueError, ArithmeticError) as exc:
            preview.append({"row": number, "status": "invalid", "detail": type(exc).__name__})
    if not confirmed:
        return {"file_sha256": file_hash, "row_count": len(rows), "ready_count": len(normalized), "preview": preview[:50], "written": False}
    run = MarketRefreshRun(mode="manual_file_import", provider_role="research_provider", requested_count=len(normalized), requested_by=actor, status="running")
    session.add(run); session.flush()
    touched: set[str] = set()
    for sector_key, bar, source in normalized:
        status = _persist_bar(session, sector_key, bar, f"manual_file_import:{source}", "research_provider")
        if status == "conflict": run.failure_count += 1
        else: run.success_count += 1; touched.add(sector_key)
        session.add(MarketRefreshItem(run_id=run.id, sector_key=sector_key, status=status, trade_date=bar.trade_date, detail=f"file_sha256={file_hash}"))
    for key in touched: _recalculate(session, key)
    run.status = "completed_with_failures" if run.failure_count else "completed"; run.finished_at = datetime.now(timezone.utc)
    session.commit()
    return {"run_id": run.id, "file_sha256": file_hash, "row_count": len(rows), "success_count": run.success_count, "failure_count": run.failure_count, "written": True}
